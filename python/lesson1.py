'''
接口自动化测试的步骤：
1、excel测试用例准备好，代码自动读取测试数据
2、发送接口请求，得到响应数据
3、断言：预期结果 vs 实际执行结果（同用例文档中）   ---通过/不通过
4、最终结果 写入excel表格中  --输出测试报告（目前课程中无，高阶班才有）

'''
import requests
import openpyxl

# 读取测试用例
def read_data(filename, sheetname):   #定义函数，设置参数
    wb = openpyxl.load_workbook(filename)
    sh = wb[sheetname]
    max_row = sh.max_row
    case_list = []
    for i in range(2,max_row+1):
        dict1 = dict(
        case_id =sh.cell(row=i,column=1).value,
        url = sh.cell(row=i,column =5).value,
        data = sh.cell(row=i,column=6).value,
        expect = sh.cell(row=i,column=7).value
        )
        case_list.append(dict1)
    return case_list   #定义返回值

#发送接口测试：
def api_fun(url,data):    #定义函数，参数
    # url_login = 'http://8.129.91.152:8766/futureloan/member/login'
    # data_login = {"mobile_phone": "13821025679","pwd": "lemon666"}
    headers = {'X-Lemonban-Media-Type':'lemonban.v2','Content-Type':'application/json'}

    result = requests.post(url=url,json=data,headers=headers).json()
    return result  #设置返回值

# 写入excel测试结果
def wirte_result(filename,sheetname,row,column,final_result):
    wb = openpyxl.load_workbook(filename)
    sh = wb[sheetname]
    sh.cell(row=row,column=column).value = final_result
    wb.save(filename)

# eval()函数 ---运行被字符串包裹的表达式
# '{"mobile_phone":"13652440101","pwd":"12345678","type":1,"reg_name":"34254sdfs"}'
# str0 = '{"mobile_phone":"13652440101","pwd":"12345678","type":1,"reg_name":"34254sdfs"}'
# dict0 = eval(str0)
# print(type(dict0))
#
# print(eval('2+3'))

# 接口测试实战,封装成函数
def execute_fun(filename,sheetname):
    cases = read_data(filename,sheetname)
    # print(cases)
    for case in cases:   #依次去访问cases中的元素，保存到定义的变量case中
        # print(case)
        case_id = case['case_id']
        url = case['url']
        data = eval(case['data'])
        # print(case_id,url,data)


        # 获取期望结果code、msg
        expect = eval(case['expect'])
        expect_code = expect['code']
        expect_msg = expect['msg']
        print('预期结果code为:{},msg为:{}'.format(expect_code, expect_msg))

        # 执行接口测试
        real_result = api_fun(url=url,data=data)
        # print(real_result)
        # # 获取实际结果code、msg
        real_code = real_result['code']
        real_msg = real_result['msg']
        print('实际结果code为:{},msg为:{}'.format(real_code, real_msg))


        # 断言：预期vs实际结果
        if real_code == expect_code and real_msg == expect_msg:
           print('这{}条测试用例执行通过！'.format(case_id))
           final_re = 'Passed'
        else:
            print('这{}条测试用例执行通过！'.format(case_id))
            final_re = 'Failed'
        print('*'*50)

 # 写入最终的测试结果到excel中
        wirte_result(filename,sheetname,case_id+1,8,final_re)


 # 调用接口自动化测试函数
# execute_fun('../test_data/test_case_api.xlsx', 'login')
